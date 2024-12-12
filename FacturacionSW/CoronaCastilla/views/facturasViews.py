from django.forms import inlineformset_factory
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Sum
from CoronaCastilla.models import Factura, Cliente, Habitacion
from CoronaCastilla.forms import FacturaForm, HabitacionForm
from django.utils import timezone
from django.contrib import messages
from datetime import datetime
import re
import os
from django.template.loader import render_to_string
from xhtml2pdf import pisa
import openpyxl
from io import BytesIO



def generate_excel(request, mes_actual, año_actual):
    meses_en_espanol = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]
    
    # Construir la ruta del archivo de plantilla
    template_path = os.path.join(os.path.dirname(__file__), '..', '..', 'FACTURAS.xlsx')

    # Verificar que el archivo existe
    if not os.path.exists(template_path):
        return HttpResponse("Error: El archivo de plantilla no se encuentra.", status=404)

    try:
        # Cargar plantilla
        wb = openpyxl.load_workbook(template_path)
    except Exception as e:
        return HttpResponse(f"Error al cargar el archivo de plantilla: {str(e)}", status=500)

    # Hoja activa
    ws = wb.active

    facturas = Factura.objects.filter(fecha_salida__year=año_actual, fecha_salida__month=mes_actual)

    # Comenzar a rellenar desde la fila 7 (por ejemplo)
    start_row = 7

    # Actualizar información en la hoja de cálculo
    ws["C3"] = f"HOSTELERÍA-EJERCICIO {meses_en_espanol[mes_actual - 1]} {año_actual}"
    for index, factura in enumerate(facturas, start=start_row):
        # Extraer cliente y NIF
        cliente_data = " ".join(factura.cliente.split())
        nif_match = re.search(r'\b\d{8}[A-Z]\b', cliente_data)
        if nif_match:
            nif = nif_match.group(0)
            # El nombre es todo lo que está antes del NIF
            nombre = cliente_data.split(nif)[0].strip()

        # Insertar datos en celdas. Por ejemplo:
        ws[f"A{index}"] = factura.numero_factura
        ws[f"B{index}"] = factura.fecha_salida.strftime('%Y-%m-%d')
        ws[f"C{index}"] = nif 
        ws[f"D{index}"] = nombre
        
        habitaciones = factura.habitaciones.all()
        habitaciones_str = ', '.join(str(habitacion) for habitacion in habitaciones)

        
        
        if factura.desayuno_precio > 0:
            ws[f"E{index}"] = habitaciones_str + ' - Desayunos'
        else:
            ws[f"E{index}"] = habitaciones_str
        ws[f"F{index}"] = factura.base_imponible
        ws[f"G{index}"] = factura.porcentaje_iva

    # Crear un archivo en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Crear nombre de archivo con mes en español
    nombre_archivo = f"facturas_{meses_en_espanol[mes_actual - 1]}_{año_actual}.xlsx"

    # Configurar la respuesta HTTP
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

    return response



def generate_pdf(factura, output_path):
    template = 'pdfFactura.html'
    habitaciones = factura.habitaciones.all()
    dias = factura.fecha_entrada - factura.fecha_salida
    alojamiento_result = sum(dias.days * habitacion.alojamiento_precio for habitacion in habitaciones)
    desayuno_result = factura.desayuno_precio
    context = {
        'factura': factura,
        'alojamiento_result': alojamiento_result,
        'desayuno_result': desayuno_result,
    }
    html = render_to_string(template, context)
    
    # Guarda el HTML para depuración
    with open('debug_output.html', 'w') as debug_file:
        debug_file.write(html)
        
    with open(output_path, "wb") as pdf_file:
        pisa_status = pisa.CreatePDF(html, dest=pdf_file)

    return pisa_status.err


def extract_name(cliente_text):
    # Encuentra el NIF en el texto usando una expresión regular
    match = re.search(r'\d{8}[A-Z]', cliente_text)
    if match:
        pos_nif = match.start()
        # Extraer el nombre del texto
        return cliente_text[:pos_nif].strip()
    return cliente_text



def view_facturas(request):
    if request.method == 'GET':
        form = FacturaForm(request.GET)  # Cargar el formulario con los datos del GET
        
        # Obtener el año desde la URL o usar el año actual si no se proporciona
        año = request.GET.get('año', datetime.now().year)

        # Extraer meses seleccionados (marcados en el checkbox)
        meses_seleccionados = request.GET.getlist('mes')


        # Si no se selecciona ningún mes, usar el mes actual por defecto
        if not meses_seleccionados:
            ahora = datetime.now()
            meses_seleccionados = [ahora.month]

        # Filtro inicial de facturas por los meses seleccionados y el año
        query = Q()
        for mes in meses_seleccionados:
            query |= Q(fecha_salida__year=año, fecha_salida__month=mes)

        facturas = Factura.objects.filter(query).order_by('-id')

        # Filtrar por nombre de cliente si está en la URL
        cliente_nombre = request.GET.get('cliente', '')
        if cliente_nombre:
            facturas = facturas.filter(cliente__icontains=cliente_nombre)

        # Total facturado
        total_facturado = facturas.aggregate(Sum('total_factura'))['total_factura__sum'] or 0

        # Extraer el nombre del cliente para cada factura (si tienes esta función)
        for factura in facturas:
            factura.cliente_nombre = extract_name(factura.cliente)

        return render(request, 'facturas.html', {
            'facturas': facturas,
            'form': form,
            'total_facturado': total_facturado,
        })
        
        
def view_factura_id(request, factura_id):
    factura = get_object_or_404(Factura, id=factura_id)
    
    # Crear el FormSet para las habitaciones
    HabitacionFormSet = inlineformset_factory(Factura, Habitacion, HabitacionForm, extra=1)
    
    if request.method == 'POST':
        form = FacturaForm(request.POST, instance=factura)
        habitacion_formset = HabitacionFormSet(request.POST, instance=factura)
        
        if form.is_valid() and habitacion_formset.is_valid():
            form.save()
            
            # Guardar cada habitación en el FormSet
            for habitacion_form in habitacion_formset:
                if habitacion_form.cleaned_data and habitacion_form not in habitacion_formset.deleted_forms:
                    habitacion = habitacion_form.save(commit=False)
                    habitacion.factura = factura  # Asignar la factura a la habitación
                    habitacion.save()
            
            # Eliminar las habitaciones marcadas para eliminar
            for habitacion_form in habitacion_formset.deleted_forms:
                habitacion_form.instance.delete()

            # Generar y guardar el PDF de la factura
            factura_name = f"factura_{str(factura.numero_factura).replace('/', '_')}.pdf"
            external_drive_path = "D:\\FACTURAS"
            if not os.path.exists("D"):
                return redirect('facturas')
            output_path = os.path.join(external_drive_path, factura_name)
            generate_pdf(factura, output_path)

            messages.success(request, "Factura actualizada correctamente.")
            return redirect('facturas')
        else:
            print('Errores de form:', form.errors)
            print('Errores de formset:', habitacion_formset.errors)
            
            
    else:
        form = FacturaForm(instance=factura)
        habitacion_formset = HabitacionFormSet(instance=factura)    
        
        if factura.numero_factura is None:
            # Calcular el siguiente número de factura
            year = timezone.now().year
            year_suffix = year % 100

            # Buscar la última factura del año
            last_invoice = Factura.objects.filter(numero_factura__endswith=f'/{year_suffix}').order_by('-numero_factura').first()
            
            if last_invoice:
                try:
                    last_number = int(last_invoice.numero_factura.split('/')[0])
                except ValueError:
                    last_number = 0
                new_number = last_number + 1
            else:
                new_number = 1

            numero_factura = f"{new_number:03}/{year_suffix}"
            form.initial['numero_factura'] = numero_factura

            
    
    # Calcular resultados de alojamiento y desayuno
    habitaciones = factura.habitaciones.all()
    print('habitaciones: ', list(habitaciones))
    dias = factura.fecha_entrada - factura.fecha_salida
    alojamiento_result = sum(dias.days * habitacion.alojamiento_precio for habitacion in habitaciones)

    context = {
        'factura': factura,
        'form': form,
        'habitacion_formset': habitacion_formset,  # Incluir el formset en el contexto
        'alojamiento_result': alojamiento_result,
        'desayuno_result': factura.desayuno_precio
    }
    
    return render(request, 'gestionarFactura.html', context)


    

def post_factura(request):
    cliente_id = request.GET.get('cliente_id')
    
    if cliente_id:
        cliente = Cliente.objects.get(id=cliente_id)
        cliente_info = f"{cliente.nombre}\n{cliente.direccion}\n{cliente.codigo_postal}\nNIF: {cliente.nif}"
    else:
        cliente_info = ''
    
    # Crear el FormSet para las habitaciones
    HabitacionFormSet = inlineformset_factory(Factura, Habitacion, form=HabitacionForm, extra=1, can_delete=True)
    
    # Calcular el siguiente número de factura
    year = timezone.now().year
    year_suffix = year % 100

    # Buscar la última factura del año y calcular el siguiente número
    last_invoice = Factura.objects.filter(numero_factura__endswith=f'/{year_suffix}').order_by('-numero_factura').first()
    if last_invoice:
        last_number = int(last_invoice.numero_factura.split('/')[0])
        new_number = last_number + 1
    else:
        new_number = 1

    numero_factura = f"{new_number:03}/{year_suffix}"
    
    if request.method == 'POST':
        form = FacturaForm(request.POST)
        habitacion_formset = HabitacionFormSet(request.POST, instance=None)  # Importante: establecer `instance=None`

        if form.is_valid() and habitacion_formset.is_valid():
            # Guardar la factura primero
            factura = form.save(commit=False)
            factura.numero_factura = None
            factura.save()

            # Asignar la factura a cada habitación antes de guardarlas
            habitaciones = habitacion_formset.save(commit=False)
            for habitacion in habitaciones:
                habitacion.factura = factura  # Asignar la factura a cada habitación
                habitacion.save()

            # Manejar la eliminación de habitaciones (si el formset tiene la opción de eliminación)
            for form_del in habitacion_formset.deleted_forms:
                if form_del.instance.pk:
                    form_del.instance.delete()

            return redirect('facturas')
        else:
            print('Formulario inválido:')
            if form.errors:
                print('errores de form:', form.errors)
            else:
                print('errores de formset:', habitacion_formset.errors)

    else:
        # Crear un nuevo formulario con el número de factura automáticamente calculado
        form = FacturaForm()
        form.fields['cliente'].initial = cliente_info
        form.fields['numero_factura'].initial = numero_factura

        # Crear un formset vacío para las habitaciones
        habitacion_formset = HabitacionFormSet(queryset=Habitacion.objects.none())

    return render(request, 'crearFactura.html', {'form': form, 'habitacion_formset': habitacion_formset})




def delete_factura(request, factura_id):
    factura = get_object_or_404(Factura, id=factura_id)
    
    if request.method == 'POST':
        factura.delete()
        return redirect('facturas')  # Redirige a la lista de facturas después de eliminar
    
    # Si no es una solicitud POST, renderiza la página de detalles de factura (o alguna otra vista)
    return render(request, 'gestionarFactura.html', {'factura': factura})


def close_factura(request, factura_id):
    factura = get_object_or_404(Factura, id=factura_id)
    
    if factura.numero_factura is None:
        # Obtener el año actual
        year = timezone.now().year
        year_suffix = year % 100
        
        # Buscar la última factura cerrada del año para asignar el siguiente número
        last_invoice = Factura.objects.filter(numero_factura__endswith=f'/{year_suffix}').order_by('-numero_factura').first()

        if last_invoice:
            last_number = int(last_invoice.numero_factura.split('/')[0])
            new_number = last_number + 1
        else:
            new_number = 1
        
        factura.numero_factura = f"{new_number:03}/{year_suffix}"
        factura.save()
        
        messages.success(request, f"Factura {factura.numero_factura} cerrada correctamente.")
        # Generar y guardar el PDF de la factura
        factura_name = f"factura_{factura.numero_factura.replace('/', '_')}.pdf"
        external_drive_path = "D:\\FACTURAS"
        if not os.path.exists("D"):
            return redirect('facturas')
        output_path = os.path.join(external_drive_path, factura_name)

        generate_pdf(factura, output_path)
       
    else:
        messages.error(request, "La factura ya está cerrada.")

    return redirect('facturas')




def close_new_factura(request):
    cliente_id = request.GET.get('cliente_id')
    
    if cliente_id:
        cliente = Cliente.objects.get(id=cliente_id)
        cliente_info = f"{cliente.nombre}\n{cliente.direccion}\n{cliente.codigo_postal}\nNIF: {cliente.nif}"
    else:
        cliente_info = ''
    
    # Crear el FormSet para las habitaciones
    HabitacionFormSet = inlineformset_factory(Factura, Habitacion, form=HabitacionForm, extra=1, can_delete=True)
    
    if request.method == 'POST':
        form = FacturaForm(request.POST)
        habitacion_formset = HabitacionFormSet(request.POST, instance=None)  # Importante: establecer `instance=None`

        if form.is_valid() and habitacion_formset.is_valid():
            # Guardar la factura primero
            factura = form.save()

            # Asignar la factura a cada habitación antes de guardarlas
            habitaciones = habitacion_formset.save(commit=False)
            for habitacion in habitaciones:
                habitacion.factura = factura  # Asignar la factura a cada habitación
                habitacion.save()

            # Manejar la eliminación de habitaciones (si el formset tiene la opción de eliminación)
            for form_del in habitacion_formset.deleted_forms:
                if form_del.instance.pk:
                    form_del.instance.delete()

            return redirect('facturas')
        else:
            print('Formulario inválido:')
            if form.errors:
                print('errores de form:', form.errors)
            else:
                print('errores de formset:', habitacion_formset.errors)

    else:
        # Crear un nuevo formulario con el número de factura automáticamente calculado
        form = FacturaForm()
        form.fields['cliente'].initial = cliente_info

        # Crear un formset vacío para las habitaciones
        habitacion_formset = HabitacionFormSet(queryset=Habitacion.objects.none())

    return redirect('facturas')